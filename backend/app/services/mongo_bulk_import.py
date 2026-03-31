from __future__ import annotations

"""Bulk import metadata into MongoDB (map-id based) + sync PostgreSQL + Neo4j.

Mục tiêu:
- Import nhiều record 1 lần (từ JSON hoặc từ file Excel).
- Dùng MAP IDs (L10 / TH10 / TH10_CD1 / TH10_CD1_B1 / TH10_CD1_B1_C1) làm "key".
- Không lưu các cột ref kiểu import_key/ref vào Mongo (chỉ dùng để resolve khi đọc Excel cũ).
- Mongo -> Postgre (auto ids) -> Neo4j (light nodes).

Mongo schema follow app/services/mongo_sync.py:
- classes:  {classID, className}
- subjects: {subjectID, classID, subjectName, subjectTitle, subjectDescription, keywordSubject, subjectUrl, subjectCategory, status, createdBy, createdAt, updatedAt}
- topics:   {topicID, subjectID, topicName, topicDescription, keywordTopic, topicUrl, topicCategory, status, createdBy, createdAt, updatedAt}
- lessons:  {lessonID, topicID, lessonName, lessonType, lessonDescription, keywordLesson, lessonUrl, lessonCategory, status, createdBy, createdAt, updatedAt}
- chunks:   {chunkID, lessonID, chunkName, chunkType, chunkUrl, keywords, chunkDescription, chunkCategory, status, createdBy, createdAt, updatedAt}

QUAN TRỌNG: chunk chỉ lưu lessonID = lesson_map (string map id), KHÔNG lưu subject/topic/class.
"""

import io
import re
from dataclasses import dataclass
from dataclasses import asdict, is_dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .mongo_client import get_mongo_client
from .keyword_embedding import embed_keyword_cached, get_keyword_embedder
from .postgre_sync_from_mongo import PgIds, sync_postgre_from_mongo_auto_ids
from .neo_sync import NeoSyncResult, sync_neo4j_from_maps_and_pg_ids
from .postgre_media_sync import sync_postgre_media_from_mongo
from .neo_media_sync import sync_media_to_neo4j


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _clean(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(v)
    s = _clean(v)
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _parse_keywords(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, list):
        return [x for x in (_clean(i) for i in v) if x]
    s = _clean(v)
    if not s:
        return []
    parts = re.split(r"[;,\n\r\t]+", s)
    return [p.strip() for p in parts if p and p.strip()]


def _uniq_keep_order_ci(values: List[Any]) -> List[str]:
    out: List[str] = []
    seen: set[str] = set()
    for item in values or []:
        s = _clean(item)
        if not s:
            continue
        key = s.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


def _active_docs(cursor) -> List[Dict[str, Any]]:
    docs: List[Dict[str, Any]] = []
    for doc in cursor:
        if str(doc.get("status") or "active").strip().lower() == "hidden":
            continue
        docs.append(doc)
    return docs


def _refresh_topic_keywords_from_lessons(*, topic_map: str, lesson_map: str = "", category: str = "document") -> Dict[str, Any] | None:
    mg = get_mongo_client()
    db = mg["db"]
    now = _now()
    category = _normalize_category(category)

    topic_q: Dict[str, Any] = {"topicID": _clean(topic_map)}
    if category:
        topic_q["topicCategory"] = category
    topic_doc = db["topics"].find_one(topic_q) if _clean(topic_map) else None

    if topic_doc is None and _clean(lesson_map):
        lesson_q: Dict[str, Any] = {"lessonID": _clean(lesson_map)}
        if category:
            lesson_q["lessonCategory"] = category
        lesson_doc = db["lessons"].find_one(lesson_q)
        if lesson_doc:
            topic_lookup: Dict[str, Any] = {"topicID": _clean(lesson_doc.get("topicID"))}
            if category:
                topic_lookup["topicCategory"] = category
            topic_doc = db["topics"].find_one(topic_lookup)

    if not topic_doc:
        return None

    topic_id = str(topic_doc.get("topicID") or "").strip()
    lesson_q = {"topicID": topic_id}
    if category:
        lesson_q["lessonCategory"] = category
    lesson_docs = _active_docs(db["lessons"].find(lesson_q).sort("lessonID", 1))

    merged: List[str] = []
    seen: set[str] = set()
    for lesson in lesson_docs:
        for kw in _parse_keywords(lesson.get("keywordLesson")):
            key = kw.casefold()
            if key in seen:
                continue
            seen.add(key)
            merged.append(kw)

    db["topics"].update_one(
        {"_id": topic_doc["_id"]},
        {"$set": {"keywordTopic": merged, "searchUpdatedAt": now, "updatedAt": now}},
    )
    return {
        "topicID": topic_id,
        "subjectID": str(topic_doc.get("subjectID") or "").strip(),
        "keywordCount": len(merged),
        "category": category,
    }


def _refresh_subject_keywords_from_topics(*, subject_map: str, category: str = "document") -> Dict[str, Any] | None:
    if not _clean(subject_map):
        return None

    mg = get_mongo_client()
    db = mg["db"]
    now = _now()
    category = _normalize_category(category)

    subject_q: Dict[str, Any] = {"subjectID": _clean(subject_map)}
    if category:
        subject_q["subjectCategory"] = category
    subject_doc = db["subjects"].find_one(subject_q)
    if not subject_doc:
        return None

    topic_q = {"subjectID": _clean(subject_map)}
    if category:
        topic_q["topicCategory"] = category
    topics = _active_docs(db["topics"].find(topic_q).sort("topicID", 1))

    merged: List[str] = []
    seen: set[str] = set()
    for topic in topics:
        for kw in _parse_keywords(topic.get("keywordTopic")):
            key = kw.casefold()
            if key in seen:
                continue
            seen.add(key)
            merged.append(kw)

    db["subjects"].update_one(
        {"_id": subject_doc["_id"]},
        {"$set": {"keywordSubject": merged, "searchUpdatedAt": now, "updatedAt": now}},
    )
    return {"subjectID": _clean(subject_map), "keywordCount": len(merged), "category": category}


def _extract_last_number(s: str) -> str:
    m = re.findall(r"\d+", s or "")
    return m[-1] if m else ""


def _derive_class_map_from_subject_map(subject_map: str) -> str:
    n = _extract_last_number(subject_map)
    return f"L{n}" if n else ""


def _parse_topic_map(topic_map: str) -> Optional[Dict[str, str]]:
    s = _clean(topic_map)
    m = re.match(r"^(.+?)_CD(\d+)$", s, flags=re.I)
    if not m:
        return None
    subject_map = m.group(1)
    topic_number = m.group(2)
    return {
        "subject_map": subject_map,
        "topic_map": s,
        "topicNumber": topic_number,
        "class_map": _derive_class_map_from_subject_map(subject_map),
    }


def _parse_lesson_map(lesson_map: str) -> Optional[Dict[str, str]]:
    s = _clean(lesson_map)
    m = re.match(r"^(.+?)_CD(\d+)_B(\d+)$", s, flags=re.I)
    if not m:
        return None
    subject_map = m.group(1)
    topic_number = m.group(2)
    lesson_number = m.group(3)
    topic_map = f"{subject_map}_CD{topic_number}"
    return {
        "subject_map": subject_map,
        "topic_map": topic_map,
        "lesson_map": s,
        "topicNumber": topic_number,
        "lessonNumber": lesson_number,
        "class_map": _derive_class_map_from_subject_map(subject_map),
    }


def _parse_chunk_map(chunk_map: str) -> Optional[Dict[str, str]]:
    s = _clean(chunk_map)
    m = re.match(r"^(.+?)_CD(\d+)_B(\d+)_C(\d+)$", s, flags=re.I)
    if not m:
        return None
    subject_map = m.group(1)
    topic_number = m.group(2)
    lesson_number = m.group(3)
    chunk_number = m.group(4)
    topic_map = f"{subject_map}_CD{topic_number}"
    lesson_map = f"{topic_map}_B{lesson_number}"
    return {
        "subject_map": subject_map,
        "topic_map": topic_map,
        "lesson_map": lesson_map,
        "chunk_map": s,
        "topicNumber": topic_number,
        "lessonNumber": lesson_number,
        "chunkNumber": chunk_number,
        "class_map": _derive_class_map_from_subject_map(subject_map),
    }


def _media_fields(media_type: str) -> Tuple[str, str, str, str, str]:
    if media_type == "image":
        return "images", "imgID", "imgName", "imgDescription", "imgUrl"
    if media_type == "video":
        return "videos", "videoID", "videoName", "videoDescription", "videoUrl"
    raise ValueError("media_type must be image or video")


def _parse_media_map(map_id: str) -> Optional[Dict[str, str]]:
    s = _clean(map_id)
    if s.upper().startswith("IMG_"):
        media_type = "image"
        follow_map = s[4:]
    elif s.upper().startswith("VD_"):
        media_type = "video"
        follow_map = s[3:]
    else:
        return None

    parsed = _parse_chunk_map(follow_map)
    if parsed:
        return {**parsed, "map_id": s, "follow_map": follow_map, "follow_type": "chunk", "media_type": media_type}

    parsed_lesson = _parse_lesson_map(follow_map)
    if parsed_lesson:
        return {**parsed_lesson, "map_id": s, "follow_map": follow_map, "follow_type": "lesson", "media_type": media_type, "chunk_map": ""}

    parsed_topic = _parse_topic_map(follow_map)
    if parsed_topic:
        return {**parsed_topic, "map_id": s, "follow_map": follow_map, "follow_type": "topic", "media_type": media_type, "lesson_map": "", "chunk_map": ""}

    if follow_map:
        return {
            "map_id": s,
            "follow_map": follow_map,
            "follow_type": "subject",
            "media_type": media_type,
            "subject_map": follow_map,
            "class_map": _derive_class_map_from_subject_map(follow_map),
            "topic_map": "",
            "lesson_map": "",
            "chunk_map": "",
        }
    return None


def _follow_id_from_media_map(map_id: str) -> Tuple[str, str]:
    parsed = _parse_media_map(map_id)
    if not parsed:
        raise ValueError(f"media map invalid: {map_id}")
    subject_map = parsed.get("subject_map") or ""
    follow_type = parsed.get("follow_type") or "subject"
    parts = [subject_map]
    topic_map = parsed.get("topic_map") or ""
    lesson_map = parsed.get("lesson_map") or ""
    chunk_map = parsed.get("chunk_map") or ""
    if follow_type in ("topic", "lesson", "chunk") and topic_map:
        topic_no = _extract_last_number(topic_map)
        if topic_no:
            parts.append(f"T{int(topic_no)}")
    if follow_type in ("lesson", "chunk") and lesson_map:
        lesson_no = _extract_last_number(lesson_map)
        if lesson_no:
            parts.append(f"L{int(lesson_no)}")
    if follow_type == "chunk" and chunk_map:
        chunk_no = _extract_last_number(chunk_map)
        if chunk_no:
            parts.append(f"C{int(chunk_no)}")
    return "_".join(parts), follow_type


def upsert_media_to_mongo(*, media_type: str, actor: str, item: Dict[str, Any], stats: Optional[Dict[str, Dict[str, int]]] = None) -> Dict[str, Any]:
    media_type = _normalize_category(media_type)
    if media_type not in ("image", "video"):
        raise ValueError("media_type must be image or video")

    parsed = _parse_media_map(_best(item, "mapID", "map_id"))
    if not parsed or parsed.get("media_type") != media_type:
        raise ValueError(f"mapID invalid for {media_type}")

    mg = get_mongo_client()
    db = mg["db"]
    now = _now()
    collection, id_field, name_field, desc_field, url_field = _media_fields(media_type)

    def _bump(col_key: str, action: str):
        if not stats:
            return
        if col_key not in stats:
            stats[col_key] = {"inserted": 0, "updated": 0}
        stats[col_key][action] = int(stats[col_key].get(action, 0)) + 1

    media_id = _best(item, id_field, "mediaID", "media_id")
    media_name = _best(item, name_field, "name")
    media_desc = _best(item, desc_field, "description")
    media_url = _best(item, url_field, "url")
    object_key = _best(item, "objectKey", "object_key")
    map_id = _best(item, "mapID", "map_id")
    status = _best(item, "status") or "active"

    existing = None
    if media_id:
        existing = db[collection].find_one({id_field: media_id})
    if existing is None and object_key:
        existing = db[collection].find_one({"objectKey": object_key})

    if existing:
        media_name = media_name or _clean(existing.get(name_field)) or media_id or map_id
        set_fields = {
            id_field: media_id or _clean(existing.get(id_field)),
            name_field: media_name,
            desc_field: media_desc or _clean(existing.get(desc_field)),
            url_field: media_url or _clean(existing.get(url_field)),
            "objectKey": object_key or _clean(existing.get("objectKey")),
            "mapID": map_id or _clean(existing.get("mapID")),
            "status": status or _clean(existing.get("status")) or "active",
            "updatedAt": now,
        }
        db[collection].update_one({"_id": existing["_id"]}, {"$set": set_fields})
        _bump(collection, "updated")
        mongo_id = str(existing["_id"])
        effective_id = set_fields[id_field]
    else:
        if not media_id:
            raise ValueError(f"{id_field} is required")
        doc = {
            id_field: media_id,
            name_field: media_name or media_id or map_id,
            desc_field: media_desc,
            url_field: media_url,
            "objectKey": object_key,
            "mapID": map_id,
            "status": status,
            "createdBy": actor or "system",
            "createdAt": now,
            "updatedAt": now,
        }
        mongo_id = str(db[collection].insert_one(doc).inserted_id)
        effective_id = media_id
        _bump(collection, "inserted")

    return {
        "collection": collection,
        "mongo_id": mongo_id,
        "media_id": effective_id,
        "map_id": map_id,
        "media_type": media_type,
        "follow_type": parsed.get("follow_type") or "subject",
        "class_map": parsed.get("class_map") or "",
        "subject_map": parsed.get("subject_map") or "",
        "topic_map": parsed.get("topic_map") or "",
        "lesson_map": parsed.get("lesson_map") or "",
        "chunk_map": parsed.get("chunk_map") or "",
    }


def _normalize_category(v: Any) -> str:
    s = _clean(v).lower()
    if s in ("image", "images"):
        return "image"
    if s in ("video", "videos"):
        return "video"
    return "document"


def _best(v: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        if not k:
            continue
        if k not in v:
            continue
        s = _clean(v.get(k))
        if s:
            return s
    return ""


@dataclass
class ImportRowResult:
    ok: bool
    level: str
    map_id: str
    mongo_id: Optional[str] = None
    postgre: Optional[PgIds] = None
    neo4j: Optional[NeoSyncResult] = None
    warning: Optional[str] = None
    error: Optional[str] = None
    row: Optional[int] = None


@dataclass
class BulkImportResult:
    ok: bool
    inserted_or_updated: Dict[str, int]
    results: List[ImportRowResult]


def upsert_chain_to_mongo_by_maps(
    *,
    level: str,
    actor: str,
    category: str,
    class_map: str,
    subject_map: str = "",
    topic_map: str = "",
    lesson_map: str = "",
    chunk_map: str = "",
    # names/fields
    class_name: str = "",
    subject_name: str = "",
    subject_title: str = "",
    subject_description: str = "",
    subject_keywords: Optional[List[str]] = None,
    subject_url: str = "",
    topic_name: str = "",
    topic_description: str = "",
    topic_keywords: Optional[List[str]] = None,
    topic_url: str = "",
    lesson_name: str = "",
    lesson_type: str = "",
    lesson_description: str = "",
    lesson_keywords: Optional[List[str]] = None,
    lesson_url: str = "",
    chunk_name: str = "",
    chunk_type: str = "",
    chunk_url: str = "",
    keywords: Optional[List[str]] = None,
    chunk_description: str = "",
    status: str = "",
    stats: Optional[Dict[str, Dict[str, int]]] = None,
) -> Dict[str, Any]:
    """Upsert 1 chain node into Mongo (class/subject/topic/lesson/chunk).

    Return: {class_id, subject_id, topic_id, lesson_id, chunk_id} (Mongo ObjectId as str)
    """

    mg = get_mongo_client()
    db = mg["db"]

    level = _clean(level).lower()
    category = _normalize_category(category)
    actor = _clean(actor) or "system"

    class_map = _clean(class_map)
    subject_map = _clean(subject_map)
    topic_map = _clean(topic_map)
    lesson_map = _clean(lesson_map)
    chunk_map = _clean(chunk_map)

    subject_kw = _uniq_keep_order_ci(subject_keywords or [])
    topic_kw = _uniq_keep_order_ci(topic_keywords or [])
    lesson_kw = _uniq_keep_order_ci(lesson_keywords or [])

    if not class_map:
        raise ValueError("class_map is required")

    now = _now()

    COL_CLASSES = "classes"
    COL_SUBJECTS = "subjects"
    COL_TOPICS = "topics"
    COL_LESSONS = "lessons"
    COL_CHUNKS = "chunks"
    COL_KEYWORDS = "keywords"

    def _bump(col_key: str, action: str):
        if not stats:
            return
        if col_key not in stats:
            stats[col_key] = {"inserted": 0, "updated": 0}
        stats[col_key][action] = int(stats[col_key].get(action, 0)) + 1

    # ===== CLASS =====
    class_filter = {"classID": class_map}
    class_doc = db[COL_CLASSES].find_one(class_filter)
    if class_doc:
        class_id = class_doc["_id"]
        db[COL_CLASSES].update_one(
            {"_id": class_id},
            {
                "$set": {
                    "className": class_name or class_doc.get("className") or class_map,
                    "updatedAt": now,
                }
            },
        )
        _bump("classes", "updated")
    else:
        class_id = db[COL_CLASSES].insert_one(
            {
                **class_filter,
                "className": class_name or class_map,
                "createdAt": now,
                "updatedAt": now,
            }
        ).inserted_id
        _bump("classes", "inserted")

    out: Dict[str, Any] = {"class_id": str(class_id)}

    # ===== SUBJECT =====
    if level in ("subject", "topic", "lesson", "chunk"):
        if not subject_map:
            raise ValueError("subject_map is required")

        subject_filter = {"subjectID": subject_map, "subjectCategory": category}
        subject_doc = db[COL_SUBJECTS].find_one(subject_filter)
        if subject_doc:
            subject_id = subject_doc["_id"]
            set_fields: Dict[str, Any] = {
                "classID": class_map,
                "subjectName": subject_name or subject_doc.get("subjectName") or subject_map,
                "subjectTitle": subject_title or subject_doc.get("subjectTitle") or "",
                "updatedAt": now,
            }
            if subject_description:
                set_fields["subjectDescription"] = subject_description
            if subject_kw:
                set_fields["keywordSubject"] = subject_kw
            if subject_url:
                set_fields["subjectUrl"] = subject_url
            db[COL_SUBJECTS].update_one({"_id": subject_id}, {"$set": set_fields})
            _bump("subjects", "updated")
        else:
            subject_id = db[COL_SUBJECTS].insert_one(
                {
                    **subject_filter,
                    "classID": class_map,
                    "subjectName": subject_name or subject_map,
                    "subjectTitle": subject_title,
                    "subjectDescription": subject_description,
                    "keywordSubject": subject_kw,
                    "subjectUrl": subject_url,
                    "status": status or "active",
                    "createdBy": actor,
                    "createdAt": now,
                    "updatedAt": now,
                }
            ).inserted_id
            _bump("subjects", "inserted")

        out["subject_id"] = str(subject_id)

    # ===== TOPIC =====
    if level in ("topic", "lesson", "chunk"):
        if not topic_map:
            raise ValueError("topic_map is required")
        topic_filter = {"topicID": topic_map, "topicCategory": category}
        topic_doc = db[COL_TOPICS].find_one(topic_filter)
        if topic_doc:
            topic_id = topic_doc["_id"]
            set_fields = {
                "subjectID": subject_map,
                "topicName": topic_name or topic_doc.get("topicName") or topic_map,
                "updatedAt": now,
            }
            if topic_description:
                set_fields["topicDescription"] = topic_description
            if topic_kw:
                set_fields["keywordTopic"] = topic_kw
            if topic_url:
                set_fields["topicUrl"] = topic_url
            db[COL_TOPICS].update_one({"_id": topic_id}, {"$set": set_fields})
            _bump("topics", "updated")
        else:
            topic_id = db[COL_TOPICS].insert_one(
                {
                    **topic_filter,
                    "subjectID": subject_map,
                    "topicName": topic_name or topic_map,
                    "topicDescription": topic_description,
                    "keywordTopic": topic_kw,
                    "topicUrl": topic_url,
                    "status": status or "active",
                    "createdBy": actor,
                    "createdAt": now,
                    "updatedAt": now,
                }
            ).inserted_id
            _bump("topics", "inserted")
        out["topic_id"] = str(topic_id)

    # ===== LESSON =====
    if level in ("lesson", "chunk"):
        if not lesson_map:
            raise ValueError("lesson_map is required")
        lesson_filter = {"lessonID": lesson_map, "lessonCategory": category}
        lesson_doc = db[COL_LESSONS].find_one(lesson_filter)
        if lesson_doc:
            lesson_id = lesson_doc["_id"]
            set_fields = {
                "topicID": topic_map,
                "lessonName": lesson_name or lesson_doc.get("lessonName") or lesson_map,
                "lessonType": lesson_type or lesson_doc.get("lessonType") or "",
                "updatedAt": now,
            }
            if lesson_description:
                set_fields["lessonDescription"] = lesson_description
            if lesson_kw:
                set_fields["keywordLesson"] = lesson_kw
            if lesson_url:
                set_fields["lessonUrl"] = lesson_url
            db[COL_LESSONS].update_one({"_id": lesson_id}, {"$set": set_fields})
            _bump("lessons", "updated")
        else:
            lesson_id = db[COL_LESSONS].insert_one(
                {
                    **lesson_filter,
                    "topicID": topic_map,
                    "lessonName": lesson_name or lesson_map,
                    "lessonType": lesson_type,
                    "lessonDescription": lesson_description,
                    "keywordLesson": lesson_kw,
                    "lessonUrl": lesson_url,
                    "status": status or "active",
                    "createdBy": actor,
                    "createdAt": now,
                    "updatedAt": now,
                }
            ).inserted_id
            _bump("lessons", "inserted")
        out["lesson_id"] = str(lesson_id)

    # ===== CHUNK =====
    if level == "chunk":
        if not chunk_map:
            raise ValueError("chunk_map is required")
        if not lesson_map:
            raise ValueError("lesson_map is required (for chunk)")
        chunk_filter = {"chunkID": chunk_map, "chunkCategory": category}
        chunk_doc = db[COL_CHUNKS].find_one(chunk_filter)
        kw = keywords or []
        if chunk_doc:
            chunk_id = chunk_doc["_id"]
            set_fields = {
                "lessonID": lesson_map,  # QUAN TRỌNG
                "chunkName": chunk_name or chunk_doc.get("chunkName") or chunk_map,
                "chunkType": chunk_type or chunk_doc.get("chunkType") or "",
                "chunkUrl": chunk_url or chunk_doc.get("chunkUrl") or "",
                "keywords": kw or chunk_doc.get("keywords") or [],
                "chunkDescription": chunk_description or chunk_doc.get("chunkDescription") or "",
                "updatedAt": now,
            }
            if status:
                set_fields["status"] = status
            db[COL_CHUNKS].update_one({"_id": chunk_id}, {"$set": set_fields})
            _bump("chunks", "updated")
        else:
            chunk_id = db[COL_CHUNKS].insert_one(
                {
                    **chunk_filter,
                    "lessonID": lesson_map,  # QUAN TRỌNG
                    "chunkName": chunk_name or chunk_map,
                    "chunkType": chunk_type,
                    "chunkUrl": chunk_url,
                    "keywords": kw,
                    "chunkDescription": chunk_description,
                    "status": status or "active",
                    "createdBy": actor,
                    "createdAt": now,
                    "updatedAt": now,
                }
            ).inserted_id
            _bump("chunks", "inserted")

        # Đồng bộ keywords collection để cùng schema với luồng import Mongo cũ.
        try:
            db[COL_KEYWORDS].delete_many({"chunkID": chunk_map})
        except Exception:
            pass

        if kw:
            embedder = get_keyword_embedder()
            for idx, kw_name in enumerate(kw, start=1):
                kw_name = _clean(kw_name)
                if not kw_name:
                    continue
                kw_map = f"{chunk_map}_K{idx}"
                db[COL_KEYWORDS].update_one(
                    {"keywordID": kw_map},
                    {
                        "$set": {
                            "chunkID": chunk_map,
                            "keywordName": kw_name,
                            "keywordEmbedding": embed_keyword_cached(kw_name),
                            "embeddingProvider": getattr(embedder, "name", ""),
                            "status": status or "active",
                            "updatedAt": now,
                        },
                        "$setOnInsert": {
                            "createdBy": actor,
                            "createdAt": now,
                        },
                    },
                    upsert=True,
                )

        out["chunk_id"] = str(chunk_id)

    return out


def _sheet_rows(ws) -> List[Tuple[int, Dict[str, Any]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if h is None else str(h).strip() for h in rows[0]]

    out: List[Tuple[int, Dict[str, Any]]] = []
    for i, r in enumerate(rows[1:], start=2):
        if r is None:
            continue
        if all((x is None or _clean(x) == "") for x in r):
            continue
        d: Dict[str, Any] = {}
        for h, v in zip(headers, r):
            if not h:
                continue
            d[h] = v
        out.append((i, d))
    return out


def parse_excel_to_payload(excel_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
    """Đọc Excel template nhiều sheet -> payload chuẩn hoá.

    Hỗ trợ 2 kiểu:
    - Template cũ: import_key + *_ref (class_ref/subject_ref/...)
    - Template mới: map IDs trực tiếp (classID/subjectID/topicID/lessonID/chunkID) và KHÔNG cần ref.
    """

    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    sheets = {name.lower(): wb[name] for name in wb.sheetnames}

    raw: Dict[str, List[Tuple[int, Dict[str, Any]]]] = {}
    for nm in ("class", "subject", "topic", "lesson", "chunk", "keyword"):
        raw[nm] = _sheet_rows(sheets[nm]) if nm in sheets else []
    raw["image"] = _sheet_rows(sheets["image"]) if "image" in sheets else (_sheet_rows(sheets["images"]) if "images" in sheets else [])
    raw["video"] = _sheet_rows(sheets["video"]) if "video" in sheets else (_sheet_rows(sheets["videos"]) if "videos" in sheets else [])

    # ===== Pass 1: build mapping import_key -> mapID (để support file cũ) =====
    class_key_to_id: Dict[str, str] = {}
    subject_key_to_id: Dict[str, str] = {}
    topic_key_to_id: Dict[str, str] = {}
    lesson_key_to_id: Dict[str, str] = {}
    chunk_key_to_id: Dict[str, str] = {}

    def _guess_class_id(row: Dict[str, Any]) -> str:
        s = _best(row, "classID", "class_id", "classMap", "class_map", "import_key")
        if re.match(r"^L\d+$", s, flags=re.I):
            return s.upper().replace("l", "L")
        name = _best(row, "className", "class_name", "class")
        n = _extract_last_number(name or s)
        return f"L{n}" if n else s

    def _guess_subject_id(row: Dict[str, Any]) -> str:
        s = _best(row, "subjectID", "subject_id", "subjectMap", "subject_map", "import_key")
        if re.match(r"^[A-Z]{1,6}\d+(?:[-_][A-Z0-9]+)?$", s, flags=re.I):
            return s
        m = re.search(r"l(10|11|12)", s, flags=re.I)
        if m:
            return f"TH{m.group(1)}"
        return s

    for _rowno, r in raw["class"]:
        k = _best(r, "import_key")
        cid = _guess_class_id(r)
        if k and cid:
            class_key_to_id[k] = cid

    for _rowno, r in raw["subject"]:
        k = _best(r, "import_key")
        sid = _guess_subject_id(r)
        if k and sid:
            subject_key_to_id[k] = sid

    for _rowno, r in raw["topic"]:
        k = _best(r, "import_key")
        tid = _best(r, "topicID", "topic_id", "topicMap", "topic_map")
        if not tid:
            subj_ref = _best(r, "subjectID", "subject_ref", "subject_map")
            subj = subject_key_to_id.get(subj_ref, subj_ref)
            tnum = _to_int(_best(r, "topic_num", "topicNum", "topicNumber"))
            if subj and tnum:
                tid = f"{subj}_CD{tnum}"
        if k and tid:
            topic_key_to_id[k] = tid

    for _rowno, r in raw["lesson"]:
        k = _best(r, "import_key")
        lid = _best(r, "lessonID", "lesson_id", "lessonMap", "lesson_map")
        if not lid:
            top_ref = _best(r, "topicID", "topic_ref", "topic_map")
            top = topic_key_to_id.get(top_ref, top_ref)
            lnum = _to_int(_best(r, "lesson_num", "lessonNum", "lessonNumber"))
            if top and lnum:
                lid = f"{top}_B{lnum}"
        if k and lid:
            lesson_key_to_id[k] = lid

    for _rowno, r in raw["chunk"]:
        k = _best(r, "import_key")
        chid = _best(r, "chunkID", "chunk_id", "chunkMap", "chunk_map")
        if not chid:
            les_ref = _best(r, "lessonID", "lesson_ref", "lesson_map")
            les = lesson_key_to_id.get(les_ref, les_ref)
            cnum = _to_int(_best(r, "chunk_label", "chunkLabel", "chunk_num", "chunkNum", "chunkNumber"))
            if les and cnum:
                chid = f"{les}_C{cnum}"
        if k and chid:
            chunk_key_to_id[k] = chid

    # ===== Keywords sheet: group by chunk_ref/chunkID =====
    keywords_by_chunk: Dict[str, List[str]] = {}
    for _rowno, r in raw["keyword"]:
        ck = _best(r, "chunkID", "chunk_ref", "chunk_map")
        ck = chunk_key_to_id.get(ck, ck)
        kw = _best(r, "keyword", "keyword_name", "keywordName")
        if ck and kw:
            keywords_by_chunk.setdefault(ck, []).append(_clean(kw))

    payload: Dict[str, List[Dict[str, Any]]] = {"classes": [], "subjects": [], "topics": [], "lessons": [], "chunks": [], "images": [], "videos": []}

    for rowno, r in raw["class"]:
        class_id = _guess_class_id(r)
        if not class_id:
            continue
        payload["classes"].append({"_row": rowno, "classID": class_id, "className": _best(r, "className", "class_name") or class_id})

    for rowno, r in raw["subject"]:
        subject_id = _guess_subject_id(r)
        if not subject_id:
            continue
        class_ref = _best(r, "classID", "class_ref", "class_map")
        class_id = class_key_to_id.get(class_ref, class_ref) or _derive_class_map_from_subject_map(subject_id)
        payload["subjects"].append(
            {
                "_row": rowno,
                "classID": class_id,
                "subjectID": subject_id,
                "subjectName": _best(r, "subjectName", "subject_name") or subject_id,
                "subjectTitle": _best(r, "subjectTitle", "subject_type", "subject_title"),
                "subjectDescription": _best(r, "subjectDescription", "subject_description", "description"),
                "keywordSubject": _parse_keywords(_best(r, "keywordSubject", "subjectKeywords", "subject_keywords")),
                "subjectCategory": _normalize_category(_best(r, "category", "subjectCategory", "subject_category")),
                "subjectUrl": _best(r, "subjectUrl", "subject_url", "url"),
            }
        )

    for rowno, r in raw["topic"]:
        topic_id = _best(r, "topicID", "topic_id", "topicMap", "topic_map")
        subject_ref = _best(r, "subjectID", "subject_ref", "subject_map")
        subject_id = subject_key_to_id.get(subject_ref, subject_ref)

        if not subject_id and topic_id:
            parsed_topic = _parse_topic_map(topic_id)
            if parsed_topic:
                subject_id = parsed_topic["subject_map"]

        if not topic_id and subject_id:
            tnum = _to_int(_best(r, "topic_num", "topicNum", "topicNumber"))
            if tnum:
                topic_id = f"{subject_id}_CD{tnum}"

        if not topic_id:
            continue

        if not subject_id:
            parsed_topic = _parse_topic_map(topic_id)
            if parsed_topic:
                subject_id = parsed_topic["subject_map"]

        if not subject_id:
            continue

        payload["topics"].append(
            {
                "_row": rowno,
                "classID": _derive_class_map_from_subject_map(subject_id),
                "subjectID": subject_id,
                "topicID": topic_id,
                "topicName": _best(r, "topicName", "topic_name") or topic_id,
                "topicDescription": _best(r, "topicDescription", "topic_description", "description"),
                "keywordTopic": _parse_keywords(_best(r, "keywordTopic", "topicKeywords", "topic_keywords")),
                "topicCategory": _normalize_category(_best(r, "category", "topicCategory", "topic_category")),
                "topicUrl": _best(r, "topicUrl", "topic_url", "url"),
            }
        )

    for rowno, r in raw["lesson"]:
        lesson_id = _best(r, "lessonID", "lesson_id", "lessonMap", "lesson_map")
        topic_ref = _best(r, "topicID", "topic_ref", "topic_map")
        topic_id = topic_key_to_id.get(topic_ref, topic_ref)

        if not topic_id and lesson_id:
            parsed_lesson = _parse_lesson_map(lesson_id)
            if parsed_lesson:
                topic_id = parsed_lesson["topic_map"]

        if not lesson_id and topic_id:
            lnum = _to_int(_best(r, "lesson_num", "lessonNum", "lessonNumber"))
            if lnum:
                lesson_id = f"{topic_id}_B{lnum}"

        if not lesson_id:
            continue

        d = _parse_lesson_map(lesson_id)
        if not d:
            continue

        payload["lessons"].append(
            {
                "_row": rowno,
                "classID": d["class_map"],
                "subjectID": d["subject_map"],
                "topicID": d["topic_map"],
                "lessonID": d["lesson_map"],
                "lessonName": _best(r, "lessonName", "lesson_name") or d["lesson_map"],
                "lessonType": _best(r, "lessonType", "lesson_type"),
                "lessonDescription": _best(r, "lessonDescription", "lesson_description", "description"),
                "keywordLesson": _parse_keywords(_best(r, "keywordLesson", "lessonKeywords", "lesson_keywords")),
                "lessonCategory": _normalize_category(_best(r, "category", "lessonCategory", "lesson_category")),
                "lessonUrl": _best(r, "lessonUrl", "lesson_url", "url"),
            }
        )

    for rowno, r in raw["chunk"]:
        chunk_id = _best(r, "chunkID", "chunk_id", "chunkMap", "chunk_map")
        if not chunk_id:
            lesson_ref = _best(r, "lessonID", "lesson_ref", "lesson_map")
            lesson_id = lesson_key_to_id.get(lesson_ref, lesson_ref)
            cnum = _to_int(_best(r, "chunk_label", "chunkLabel", "chunk_num", "chunkNum", "chunkNumber"))
            if lesson_id and cnum:
                chunk_id = f"{lesson_id}_C{cnum}"
        if not chunk_id:
            continue
        d = _parse_chunk_map(chunk_id)
        if not d:
            continue

        kws = _parse_keywords(_best(r, "keywords", "keyword"))
        kws += keywords_by_chunk.get(chunk_id, [])
        # unique keep order
        seen: set[str] = set()
        kws_u: List[str] = []
        for kw in kws:
            kwc = _clean(kw)
            if not kwc or kwc in seen:
                continue
            seen.add(kwc)
            kws_u.append(kwc)

        payload["chunks"].append(
            {
                "_row": rowno,
                "classID": d["class_map"],
                "subjectID": d["subject_map"],
                "topicID": d["topic_map"],
                "lessonID": d["lesson_map"],
                "chunkID": d["chunk_map"],
                "chunkName": _best(r, "chunkName", "chunk_name") or d["chunk_map"],
                "chunkType": _best(r, "chunkType", "chunk_type"),
                "chunkCategory": _normalize_category(_best(r, "category", "chunkCategory", "chunk_category")),
                "chunkUrl": _best(r, "chunkUrl", "chunk_url", "url"),
                "keywords": kws_u,
                "chunkDescription": _best(r, "chunkDescription", "chunk_des", "chunk_description"),
                "status": _best(r, "status"),
            }
        )

    for rowno, r in raw["image"]:
        map_id = _best(r, "mapID", "map_id")
        parsed = _parse_media_map(map_id)
        if not parsed or parsed.get("media_type") != "image":
            continue
        payload["images"].append(
            {
                "_row": rowno,
                "imgID": _best(r, "imgID", "img_id"),
                "imgName": _best(r, "imgName", "img_name", "name"),
                "imgDescription": _best(r, "imgDescription", "img_description", "description"),
                "imgUrl": _best(r, "imgUrl", "img_url", "url"),
                "objectKey": _best(r, "objectKey", "object_key"),
                "mapID": map_id,
                "category": "image",
                "classID": parsed.get("class_map") or "",
                "subjectID": parsed.get("subject_map") or "",
                "topicID": parsed.get("topic_map") or "",
                "lessonID": parsed.get("lesson_map") or "",
                "chunkID": parsed.get("chunk_map") or "",
            }
        )

    for rowno, r in raw["video"]:
        map_id = _best(r, "mapID", "map_id")
        parsed = _parse_media_map(map_id)
        if not parsed or parsed.get("media_type") != "video":
            continue
        payload["videos"].append(
            {
                "_row": rowno,
                "videoID": _best(r, "videoID", "video_id"),
                "videoName": _best(r, "videoName", "video_name", "name"),
                "videoDescription": _best(r, "videoDescription", "video_description", "description"),
                "videoUrl": _best(r, "videoUrl", "video_url", "url"),
                "objectKey": _best(r, "objectKey", "object_key"),
                "mapID": map_id,
                "category": "video",
                "classID": parsed.get("class_map") or "",
                "subjectID": parsed.get("subject_map") or "",
                "topicID": parsed.get("topic_map") or "",
                "lessonID": parsed.get("lesson_map") or "",
                "chunkID": parsed.get("chunk_map") or "",
            }
        )

    return payload


def bulk_import_payload(
    *,
    payload: Dict[str, List[Dict[str, Any]]],
    actor: str,
    sync_postgre: bool = True,
    sync_neo4j: bool = True,
    mongo_stats: Optional[Dict[str, Dict[str, int]]] = None,
) -> BulkImportResult:
    """Import payload into Mongo, then sync PG + Neo."""

    actor = _clean(actor) or "system"
    inserted_or_updated = {"classes": 0, "subjects": 0, "topics": 0, "lessons": 0, "chunks": 0, "images": 0, "videos": 0}
    results: List[ImportRowResult] = []
    dirty_topics: set[Tuple[str, str]] = set()
    dirty_subjects: set[Tuple[str, str]] = set()

    def _sync_media_row(*, media_type: str, mongo_id: str, media_id: str, map_id: str) -> Tuple[Any, Any, Optional[str]]:
        pg_res = None
        neo_res = None
        warning = None
        follow_id = ""
        follow_type = ""
        if sync_postgre:
            pg_res = sync_postgre_media_from_mongo(media_type=media_type, mongo_id=mongo_id)
            follow_id = getattr(pg_res, "follow_id", "") or ""
            follow_type = getattr(pg_res, "follow_type", "") or ""
            media_id = getattr(pg_res, "media_id", media_id) or media_id
        else:
            follow_id, follow_type = _follow_id_from_media_map(map_id)
        if sync_neo4j:
            try:
                neo_res = sync_media_to_neo4j(
                    media_type=media_type,
                    media_id=media_id,
                    mongo_id=mongo_id,
                    follow_id=follow_id,
                    follow_type=follow_type,
                )
            except Exception as exc:
                warning = f"Neo4j sync failed: {exc}"
        return pg_res, neo_res, warning

    def _import_level(level: str, items: List[Dict[str, Any]]):
        nonlocal inserted_or_updated, results
        for it in items or []:
            rowno = it.get("_row") if isinstance(it, dict) else None
            try:
                category = _normalize_category(
                    it.get("chunkCategory")
                    or it.get("lessonCategory")
                    or it.get("topicCategory")
                    or it.get("subjectCategory")
                    or it.get("category")
                    or "document"
                )

                if level in ("image", "video"):
                    media_res = upsert_media_to_mongo(media_type=level, actor=actor, item=it, stats=mongo_stats)
                    inserted_or_updated[f"{level}s"] += 1
                    pg_media = None
                    neo_media = None
                    warning = None
                    if sync_postgre or sync_neo4j:
                        try:
                            pg_media, neo_media, warning = _sync_media_row(
                                media_type=level,
                                mongo_id=media_res["mongo_id"],
                                media_id=media_res["media_id"],
                                map_id=media_res["map_id"],
                            )
                        except Exception as exc:
                            warning = f"Sync failed: {exc}"
                    results.append(
                        ImportRowResult(
                            ok=True,
                            level=level,
                            map_id=media_res["map_id"],
                            mongo_id=media_res["mongo_id"],
                            postgre=pg_media,
                            neo4j=neo_media,
                            warning=warning,
                            row=rowno,
                        )
                    )
                    continue

                class_map = _best(it, "classID", "class_map", "classMap")
                subject_map = _best(it, "subjectID", "subject_map", "subjectMap")
                topic_map = _best(it, "topicID", "topic_map", "topicMap")
                lesson_map = _best(it, "lessonID", "lesson_map", "lessonMap")
                chunk_map = _best(it, "chunkID", "chunk_map", "chunkMap")

                # Auto-derive chain maps nếu user chỉ nhập map sâu nhất
                if level == "chunk" and chunk_map:
                    d = _parse_chunk_map(chunk_map)
                    if d:
                        class_map = class_map or d["class_map"]
                        subject_map = subject_map or d["subject_map"]
                        topic_map = topic_map or d["topic_map"]
                        lesson_map = lesson_map or d["lesson_map"]
                elif level == "lesson" and lesson_map:
                    d = _parse_lesson_map(lesson_map)
                    if d:
                        class_map = class_map or d["class_map"]
                        subject_map = subject_map or d["subject_map"]
                        topic_map = topic_map or d["topic_map"]
                elif level == "topic" and topic_map:
                    d = _parse_topic_map(topic_map)
                    if d:
                        class_map = class_map or d["class_map"]
                        subject_map = subject_map or d["subject_map"]
                elif level == "subject" and subject_map:
                    class_map = class_map or _derive_class_map_from_subject_map(subject_map)

                mongo_ids = upsert_chain_to_mongo_by_maps(
                    level=level,
                    actor=actor,
                    category=category,
                    class_map=class_map,
                    subject_map=subject_map,
                    topic_map=topic_map,
                    lesson_map=lesson_map,
                    chunk_map=chunk_map,
                    class_name=_best(it, "className", "class_name"),
                    subject_name=_best(it, "subjectName", "subject_name"),
                    subject_title=_best(it, "subjectTitle", "subject_title"),
                    subject_description=_best(it, "subjectDescription", "subject_description"),
                    subject_keywords=it.get("keywordSubject") if isinstance(it.get("keywordSubject"), list) else _parse_keywords(it.get("keywordSubject")),
                    subject_url=_best(it, "subjectUrl", "subject_url"),
                    topic_name=_best(it, "topicName", "topic_name"),
                    topic_description=_best(it, "topicDescription", "topic_description"),
                    topic_keywords=it.get("keywordTopic") if isinstance(it.get("keywordTopic"), list) else _parse_keywords(it.get("keywordTopic")),
                    topic_url=_best(it, "topicUrl", "topic_url"),
                    lesson_name=_best(it, "lessonName", "lesson_name"),
                    lesson_type=_best(it, "lessonType", "lesson_type"),
                    lesson_description=_best(it, "lessonDescription", "lesson_description"),
                    lesson_keywords=it.get("keywordLesson") if isinstance(it.get("keywordLesson"), list) else _parse_keywords(it.get("keywordLesson")),
                    lesson_url=_best(it, "lessonUrl", "lesson_url"),
                    chunk_name=_best(it, "chunkName", "chunk_name"),
                    chunk_type=_best(it, "chunkType", "chunk_type"),
                    chunk_url=_best(it, "chunkUrl", "chunk_url"),
                    keywords=it.get("keywords") if isinstance(it.get("keywords"), list) else _parse_keywords(it.get("keywords")),
                    chunk_description=_best(it, "chunkDescription", "chunk_description"),
                    status=_best(it, "status"),
                    stats=mongo_stats,
                )

                inserted_or_updated[f"{level}s" if level != "class" else "classes"] += 1
                map_id = chunk_map or lesson_map or topic_map or subject_map or class_map

                results.append(
                    ImportRowResult(
                        ok=True,
                        level=level,
                        map_id=map_id,
                        mongo_id=mongo_ids.get(f"{level}_id"),
                        row=rowno,
                    )
                )

                if level == "lesson" and lesson_map:
                    d = _parse_lesson_map(lesson_map)
                    if d:
                        dirty_topics.add((d["topic_map"], category))
                        dirty_subjects.add((d["subject_map"], category))
            except Exception as e:
                map_id = (
                    _best(it, "chunkID", "chunk_map")
                    or _best(it, "lessonID", "lesson_map")
                    or _best(it, "topicID", "topic_map")
                    or _best(it, "subjectID", "subject_map")
                    or _best(it, "classID", "class_map")
                    or ""
                )
                results.append(ImportRowResult(ok=False, level=level, map_id=map_id, error=str(e), row=rowno))

    # order
    _import_level("class", payload.get("classes") or [])
    _import_level("subject", payload.get("subjects") or [])
    _import_level("topic", payload.get("topics") or [])
    _import_level("lesson", payload.get("lessons") or [])
    _import_level("chunk", payload.get("chunks") or [])
    _import_level("image", payload.get("images") or [])
    _import_level("video", payload.get("videos") or [])

    # Lesson keyword import theo Excel phải được đẩy lên topic -> subject như luồng upload tay.
    refresh_warnings: List[str] = []
    for topic_map, category in sorted(dirty_topics):
        try:
            topic_res = _refresh_topic_keywords_from_lessons(topic_map=topic_map, category=category)
            if topic_res and topic_res.get("subjectID"):
                dirty_subjects.add((str(topic_res["subjectID"]), category))
        except Exception as exc:
            refresh_warnings.append(f"topic:{topic_map}:{exc}")

    for subject_map, category in sorted(dirty_subjects):
        try:
            _refresh_subject_keywords_from_topics(subject_map=subject_map, category=category)
        except Exception as exc:
            refresh_warnings.append(f"subject:{subject_map}:{exc}")

    # ===== Build unique sync tasks (prefer deepest) =====
    chunks: set[str] = set()
    lessons: set[str] = set()
    topics: set[str] = set()
    subjects: set[str] = set()
    classes: set[str] = set()

    for r in results:
        if not r.ok:
            continue
        if r.level == "chunk" and r.map_id:
            chunks.add(r.map_id)
        elif r.level == "lesson" and r.map_id:
            lessons.add(r.map_id)
        elif r.level == "topic" and r.map_id:
            topics.add(r.map_id)
        elif r.level == "subject" and r.map_id:
            subjects.add(r.map_id)
        elif r.level == "class" and r.map_id:
            classes.add(r.map_id)

    # remove covered shallow tasks
    lessons = {l for l in lessons if not any(c.startswith(f"{l}_C") for c in chunks)}
    topics = {t for t in topics if not any(c.startswith(f"{t}_B") for c in chunks) and not any(l.startswith(f"{t}_B") for l in lessons)}
    subjects = {
        s
        for s in subjects
        if not any(c.startswith(f"{s}_CD") for c in chunks)
        and not any(l.startswith(f"{s}_CD") for l in lessons)
        and not any(t.startswith(f"{s}_CD") for t in topics)
    }

    first_idx: Dict[Tuple[str, str], int] = {}
    for idx, r in enumerate(results):
        if not r.ok:
            continue
        key = (r.level, r.map_id)
        if key not in first_idx:
            first_idx[key] = idx

    if refresh_warnings:
        for idx, r in enumerate(results):
            if r.ok:
                existing = r.warning or ""
                joined = "; ".join(refresh_warnings)
                results[idx].warning = f"{existing}; {joined}".strip("; ").strip()
                break

    def _attach(level: str, map_id: str, pg_ids: Optional[PgIds], neo_res: Optional[NeoSyncResult], warning: Optional[str]):
        idx = first_idx.get((level, map_id))
        if idx is None:
            return
        if pg_ids is not None:
            results[idx].postgre = pg_ids
        if neo_res is not None:
            results[idx].neo4j = neo_res
        if warning:
            results[idx].warning = warning

    def _run_sync(level: str, map_id: str) -> Tuple[Optional[PgIds], Optional[NeoSyncResult], Optional[str]]:
        if not sync_postgre and not sync_neo4j:
            return None, None, None

        class_map = ""
        subject_map = ""
        topic_map = ""
        lesson_map = ""
        chunk_map = ""

        if level == "chunk":
            d = _parse_chunk_map(map_id)
            if not d:
                raise ValueError(f"chunk map invalid: {map_id}")
            class_map, subject_map, topic_map, lesson_map, chunk_map = (
                d["class_map"],
                d["subject_map"],
                d["topic_map"],
                d["lesson_map"],
                d["chunk_map"],
            )
        elif level == "lesson":
            d = _parse_lesson_map(map_id)
            if not d:
                raise ValueError(f"lesson map invalid: {map_id}")
            class_map, subject_map, topic_map, lesson_map = d["class_map"], d["subject_map"], d["topic_map"], d["lesson_map"]
        elif level == "topic":
            d = _parse_topic_map(map_id)
            if not d:
                raise ValueError(f"topic map invalid: {map_id}")
            class_map, subject_map, topic_map = d["class_map"], d["subject_map"], d["topic_map"]
        elif level == "subject":
            subject_map = map_id
            class_map = _derive_class_map_from_subject_map(subject_map)
        elif level == "class":
            class_map = map_id
        else:
            raise ValueError(f"Unknown level: {level}")

        pg_ids: Optional[PgIds] = None
        neo_res: Optional[NeoSyncResult] = None
        warning: Optional[str] = None

        if sync_postgre:
            pg_ids = sync_postgre_from_mongo_auto_ids(
                class_map=class_map,
                subject_map=subject_map,
                topic_map=topic_map,
                lesson_map=lesson_map,
                chunk_map=chunk_map,
            )

        if sync_neo4j and pg_ids is not None:
            try:
                neo_res = sync_neo4j_from_maps_and_pg_ids(
                    class_map=class_map,
                    subject_map=subject_map,
                    topic_map=topic_map,
                    lesson_map=lesson_map,
                    chunk_map=chunk_map,
                    pg_ids=pg_ids,
                    actor=actor,
                )
            except Exception as e:
                warning = f"Neo4j sync failed: {e}"

        return pg_ids, neo_res, warning

    # run sync tasks (deep -> shallow)
    for ch in sorted(chunks):
        try:
            pg_ids, neo_res, warning = _run_sync("chunk", ch)
            _attach("chunk", ch, pg_ids, neo_res, warning)
        except Exception as e:
            # hide chunk if PG sync fails
            try:
                mg = get_mongo_client()
                mg["db"]["chunks"].update_one({"chunkID": ch}, {"$set": {"status": "hidden", "updatedAt": _now()}})
            except Exception:
                pass
            idx = first_idx.get(("chunk", ch))
            if idx is not None:
                results[idx].ok = False
                results[idx].error = f"Sync failed: {e}"

    for les in sorted(lessons):
        try:
            pg_ids, neo_res, warning = _run_sync("lesson", les)
            _attach("lesson", les, pg_ids, neo_res, warning)
        except Exception as e:
            idx = first_idx.get(("lesson", les))
            if idx is not None:
                results[idx].warning = f"Sync failed: {e}"

    for tp in sorted(topics):
        try:
            pg_ids, neo_res, warning = _run_sync("topic", tp)
            _attach("topic", tp, pg_ids, neo_res, warning)
        except Exception as e:
            idx = first_idx.get(("topic", tp))
            if idx is not None:
                results[idx].warning = f"Sync failed: {e}"

    for sb in sorted(subjects):
        try:
            pg_ids, neo_res, warning = _run_sync("subject", sb)
            _attach("subject", sb, pg_ids, neo_res, warning)
        except Exception as e:
            idx = first_idx.get(("subject", sb))
            if idx is not None:
                results[idx].warning = f"Sync failed: {e}"

    for cl in sorted(classes):
        try:
            pg_ids, neo_res, warning = _run_sync("class", cl)
            _attach("class", cl, pg_ids, neo_res, warning)
        except Exception as e:
            idx = first_idx.get(("class", cl))
            if idx is not None:
                results[idx].warning = f"Sync failed: {e}"

    ok = all(r.ok for r in results) if results else True
    return BulkImportResult(ok=ok, inserted_or_updated=inserted_or_updated, results=results)


# ========================= Public helper for router =========================


def _jsonify(obj: Any) -> Any:
    """Convert dataclasses + datetime to JSON-serializable structures."""
    if obj is None:
        return None
    if isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, list):
        return [_jsonify(x) for x in obj]
    if isinstance(obj, tuple):
        return [_jsonify(x) for x in obj]
    if isinstance(obj, dict):
        return {str(k): _jsonify(v) for k, v in obj.items()}
    if is_dataclass(obj):
        return _jsonify(asdict(obj))
    return str(obj)


def import_metadata_xlsx_bytes(
    excel_bytes: bytes,
    *,
    actor: str,
    category: str = "document",
    do_sync: bool = True,
    sync_postgre: Optional[bool] = None,
    sync_neo4j: Optional[bool] = None,
) -> Dict[str, Any]:
    """Router-friendly API: Excel bytes -> Mongo upsert -> (optional) sync PG + Neo.

    - actor: user name
    - category: fallback category if rows don't provide *_Category
    - do_sync: legacy flag used by router. If False => no PG/Neo sync.

    Return a JSON-serializable dict.
    """

    if excel_bytes is None or len(excel_bytes) == 0:
        raise ValueError("Empty excel content")

    # Resolve sync flags
    if not do_sync:
        sync_postgre = False
        sync_neo4j = False
    else:
        if sync_postgre is None:
            sync_postgre = True
        if sync_neo4j is None:
            sync_neo4j = True

    payload = parse_excel_to_payload(excel_bytes)

    # Apply fallback category if row doesn't have one
    fallback = _normalize_category(category)
    for it in payload.get("subjects", []):
        if not _clean(it.get("subjectCategory")):
            it["subjectCategory"] = fallback
    for it in payload.get("topics", []):
        if not _clean(it.get("topicCategory")):
            it["topicCategory"] = fallback
    for it in payload.get("lessons", []):
        if not _clean(it.get("lessonCategory")):
            it["lessonCategory"] = fallback
    for it in payload.get("chunks", []):
        if not _clean(it.get("chunkCategory")):
            it["chunkCategory"] = fallback
    for it in payload.get("images", []):
        if not _clean(it.get("category")):
            it["category"] = "image"
    for it in payload.get("videos", []):
        if not _clean(it.get("category")):
            it["category"] = "video"

    mongo_stats: Dict[str, Dict[str, int]] = {
        "classes": {"inserted": 0, "updated": 0},
        "subjects": {"inserted": 0, "updated": 0},
        "topics": {"inserted": 0, "updated": 0},
        "lessons": {"inserted": 0, "updated": 0},
        "chunks": {"inserted": 0, "updated": 0},
        "images": {"inserted": 0, "updated": 0},
        "videos": {"inserted": 0, "updated": 0},
    }

    res = bulk_import_payload(
        payload=payload,
        actor=actor,
        sync_postgre=bool(sync_postgre),
        sync_neo4j=bool(sync_neo4j),
        mongo_stats=mongo_stats,
    )

    # Build UI-friendly summary
    errors: List[Dict[str, Any]] = []
    failed = 0
    for r in res.results or []:
        if not r.ok:
            failed += 1
            errors.append(
                {
                    "row": r.row,
                    "level": r.level,
                    "map_id": r.map_id,
                    "error": r.error,
                }
            )
        elif r.warning:
            errors.append(
                {
                    "row": r.row,
                    "level": r.level,
                    "map_id": r.map_id,
                    "warning": r.warning,
                }
            )

    sync_obj: Dict[str, Any]
    if not do_sync:
        sync_obj = {"ok": True, "failed": 0, "skipped": True}
    else:
        sync_obj = {"ok": res.ok, "failed": failed, "skipped": False}

    return {
        "mongo": _jsonify(mongo_stats),
        "sync": _jsonify(sync_obj),
        "errors": _jsonify(errors),
        # Keep full details for debugging (frontend can ignore)
        "detail": _jsonify(res),
    }
